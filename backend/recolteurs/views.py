"""
recolteurs/views.py — API REST pour la gestion du personnel (recolteurs).

Expose PersonnelViewSet (ModelViewSet) :
  GET/POST         /api/recolteurs/              → liste et creation
  GET/PATCH/DELETE /api/recolteurs/:id/          → detail, modification, suppression
  GET              /api/recolteurs/stats/         → statistiques annuelles par recolteur
  GET              /api/recolteurs/:id/analytics/ → analytiques detaillees d'un recolteur
  GET              /api/recolteurs/export/        → export Excel du personnel

Le endpoint stats/ annote chaque recolteur avec :
  grands/moyens/petits — regimes recoltes par type pour l'annee filtree
  total_regimes        — total tous types confondus
  fiches_count         — nombre de fiches distinctes auxquelles il a participe
  last_recolte         — date de sa derniere participation validee

Le endpoint analytics/:id calcule egalement le rang du recolteur parmi
tous les recolteurs actifs de l'annee (global ou filtre par superviseur).
"""
import io
from decimal import Decimal

from django.http import HttpResponse
from django.db.models import Sum, Count, Max, Q
from django.db.models.functions import Coalesce
from django.db.models import Prefetch
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response

from recoltes.models import FicheRecolte, FicheRecolteDetail, FicheRecolteLigne
from .models import Personnel
from .serializers import PersonnelSerializer


def _xlsx_response(filename: str) -> tuple:
    wb = Workbook()
    ws = wb.active
    return wb, ws, filename


def _style_header(ws, headers: list, fill_color="1F4E79"):
    fill = PatternFill(fill_type="solid", fgColor=fill_color)
    font = Font(bold=True, color="FFFFFF")
    align = Alignment(horizontal="center")
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = font
        cell.fill = fill
        cell.alignment = align
    ws.auto_filter.ref = ws.dimensions


def _send_wb(wb, filename: str) -> HttpResponse:
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


from utils.audit import log_action, snapshot as audit_snapshot, diff_fields, build_revert_meta

_PERSONNEL_LABELS = {
    "nom":              "Nom",
    "lieu_residence":   "Lieu de résidence",
    "numero_telephone": "Téléphone",
    "whatsapp_actif":   "WhatsApp actif",
    "est_wave":         "Wave",
    "date_naissance":   "Date de naissance",
}


class PersonnelViewSet(viewsets.ModelViewSet):
    """CRUD complet pour le personnel (anciennement Recolteur)."""
    queryset = Personnel.objects.all().order_by("-id")
    serializer_class = PersonnelSerializer

    def create(self, request, *args, **kwargs):
        response = super().create(request, *args, **kwargs)
        if response.status_code in (200, 201):
            try:
                inst = Personnel.objects.get(pk=response.data["id"])
                snap = audit_snapshot(inst, _PERSONNEL_LABELS)
            except Exception:
                snap = {}
            nom = response.data.get("nom", "")
            log_action(request.user, "creation_recolteur",
                       detail=f"Récolteur « {nom} » créé.",
                       meta={"snapshot": snap})
        return response

    def partial_update(self, request, *args, **kwargs):
        instance = self.get_object()
        pk = instance.pk
        before = audit_snapshot(instance, _PERSONNEL_LABELS)
        before_raw = build_revert_meta(instance, _PERSONNEL_LABELS)
        response = super().partial_update(request, *args, **kwargs)
        if response.status_code == 200:
            fresh = Personnel.objects.get(pk=pk)
            after = audit_snapshot(fresh, _PERSONNEL_LABELS)
            changes = diff_fields(before, after)
            log_action(request.user, "modification_recolteur",
                       detail=f"Récolteur « {fresh.nom} » modifié.",
                       meta={"changes": changes, "object_id": pk, "object_type": "recolteur", "before_raw": before_raw})
        return response

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        snap = audit_snapshot(instance, _PERSONNEL_LABELS)
        before_raw = build_revert_meta(instance, _PERSONNEL_LABELS)
        before_raw["id"] = instance.pk
        # IDs exacts des lignes liées — utilisés sur le revert pour éviter toute ambiguïté par nom
        before_raw["ligne_ids"] = list(
            instance.lignes_recolte.values_list("id", flat=True)
        )
        log_action(request.user, "suppression_recolteur",
                   detail=f"Récolteur « {instance.nom} » supprimé.",
                   meta={"snapshot": snap, "object_type": "recolteur", "before_raw": before_raw})
        return super().destroy(request, *args, **kwargs)

    @action(detail=False, methods=["get"], url_path="stats")
    def stats(self, request):
        today = timezone.now().date()
        year = int(request.query_params.get("year", today.year))

        personnel = (
            Personnel.objects.annotate(
                grands=Coalesce(
                    Sum("lignes_recolte__details__quantite",
                        filter=Q(lignes_recolte__fiche__date__year=year, lignes_recolte__fiche__statut="valide", lignes_recolte__regime_type="grands")), 0),
                moyens=Coalesce(
                    Sum("lignes_recolte__details__quantite",
                        filter=Q(lignes_recolte__fiche__date__year=year, lignes_recolte__fiche__statut="valide", lignes_recolte__regime_type="moyens")), 0),
                petits=Coalesce(
                    Sum("lignes_recolte__details__quantite",
                        filter=Q(lignes_recolte__fiche__date__year=year, lignes_recolte__fiche__statut="valide", lignes_recolte__regime_type="petits")), 0),
                total_regimes=Coalesce(
                    Sum("lignes_recolte__details__quantite",
                        filter=Q(lignes_recolte__fiche__date__year=year, lignes_recolte__fiche__statut="valide")), 0),
                fiches_count=Count("lignes_recolte__fiche", distinct=True,
                                   filter=Q(lignes_recolte__fiche__date__year=year, lignes_recolte__fiche__statut="valide")),
                last_recolte=Max("lignes_recolte__fiche__date",
                                 filter=Q(lignes_recolte__fiche__date__year=year, lignes_recolte__fiche__statut="valide")),
            )
            .values("id", "numero_telephone", "nom", "lieu_residence",
                    "grands", "moyens", "petits", "total_regimes", "fiches_count", "last_recolte")
            .order_by("-total_regimes", "nom")
        )
        return Response({"year": year, "recolteurs": list(personnel)})

    @action(detail=True, methods=["get"], url_path="analytics")
    def analytics(self, request, pk=None):
        from django.contrib.auth.models import User as DjangoUser
        personne = self.get_object()
        today = timezone.now().date()
        year = int(request.query_params.get("year", today.year))
        prev_year = year - 1

        # Filtre optionnel par superviseur
        created_by_id = request.query_params.get("created_by")
        cb_user = None
        if created_by_id:
            try:
                cb_user = DjangoUser.objects.get(pk=int(created_by_id))
            except (ValueError, DjangoUser.DoesNotExist):
                pass
        detail_extra = {"ligne__fiche__created_by": cb_user} if cb_user else {}
        fiche_extra  = {"created_by": cb_user} if cb_user else {}
        ligne_extra  = {"fiche__created_by": cb_user} if cb_user else {}

        month_labels = ["Jan", "Fev", "Mar", "Avr", "Mai", "Juin", "Juil", "Aout", "Sept", "Oct", "Nov", "Dec"]

        def monthly_totals(target_year):
            qs = (
                FicheRecolteDetail.objects.filter(
                    ligne__recolteur=personne, ligne__fiche__date__year=target_year,
                    ligne__fiche__statut="valide", **detail_extra
                )
                .values("ligne__fiche__date__month")
                .annotate(total=Sum("quantite"))
            )
            by_month = {row["ligne__fiche__date__month"]: int(row["total"] or 0) for row in qs}
            return {"labels": month_labels, "data": [by_month.get(m, 0) for m in range(1, 13)]}

        start_year = year - 4
        yearly_qs = (
            FicheRecolteDetail.objects.filter(
                ligne__recolteur=personne, ligne__fiche__date__year__gte=start_year,
                ligne__fiche__statut="valide", **detail_extra
            )
            .values("ligne__fiche__date__year")
            .annotate(total=Sum("quantite"))
            .order_by("ligne__fiche__date__year")
        )
        yearly_map = {r["ligne__fiche__date__year"]: int(r["total"] or 0) for r in yearly_qs}
        yearly_labels = list(range(start_year, year + 1))

        yearly_regime_qs = (
            FicheRecolteDetail.objects.filter(
                ligne__recolteur=personne, ligne__fiche__date__year__gte=start_year,
                ligne__fiche__statut="valide", **detail_extra
            )
            .values("ligne__fiche__date__year", "ligne__regime_type")
            .annotate(total=Sum("quantite"))
        )
        grands_by_year, moyens_by_year, petits_by_year = {}, {}, {}
        for r in yearly_regime_qs:
            yr = r["ligne__fiche__date__year"]
            t = int(r["total"] or 0)
            rt = r["ligne__regime_type"]
            if rt == "grands":
                grands_by_year[yr] = grands_by_year.get(yr, 0) + t
            elif rt == "moyens":
                moyens_by_year[yr] = moyens_by_year.get(yr, 0) + t
            elif rt == "petits":
                petits_by_year[yr] = petits_by_year.get(yr, 0) + t

        cb_q = Q(lignes_recolte__fiche__created_by=cb_user) if cb_user else Q()
        valide_q = Q(lignes_recolte__fiche__statut="valide")
        base = (
            Personnel.objects.filter(id=personne.id)
            .annotate(
                grands=Coalesce(Sum("lignes_recolte__details__quantite",
                                    filter=Q(lignes_recolte__fiche__date__year=year, lignes_recolte__regime_type="grands") & cb_q & valide_q), 0),
                moyens=Coalesce(Sum("lignes_recolte__details__quantite",
                                    filter=Q(lignes_recolte__fiche__date__year=year, lignes_recolte__regime_type="moyens") & cb_q & valide_q), 0),
                petits=Coalesce(Sum("lignes_recolte__details__quantite",
                                    filter=Q(lignes_recolte__fiche__date__year=year, lignes_recolte__regime_type="petits") & cb_q & valide_q), 0),
                total_regimes=Coalesce(Sum("lignes_recolte__details__quantite",
                                           filter=Q(lignes_recolte__fiche__date__year=year) & cb_q & valide_q), 0),
            )
            .values("grands", "moyens", "petits", "total_regimes")
            .first() or {}
        )

        salaire_total = float(
            FicheRecolteLigne.objects.filter(recolteur=personne, fiche__date__year=year, fiche__statut="valide", **ligne_extra)
            .aggregate(t=Coalesce(Sum("salaire_calcule"), Decimal("0")))["t"] or 0
        )

        # Rang du recolteur parmi tous les actifs de l'annee.
        # Si un filtre superviseur est actif, le rang est calcule dans le perimetre
        # de ce superviseur uniquement (pas le rang global de la palmeraie).
        all_totals = (
            FicheRecolteDetail.objects.filter(ligne__fiche__date__year=year, ligne__fiche__statut="valide", **detail_extra)
            .exclude(ligne__recolteur__isnull=True)
            .values("ligne__recolteur")
            .annotate(total=Sum("quantite"))
            .order_by("-total")
        )
        rang = next((i + 1 for i, r in enumerate(all_totals) if r["ligne__recolteur"] == personne.id), None)
        nb_recolteurs_actifs = all_totals.count()

        secteurs_qs = (
            FicheRecolteDetail.objects.filter(
                ligne__recolteur=personne, ligne__fiche__date__year=year,
                ligne__fiche__statut="valide", **detail_extra
            )
            .exclude(secteur__isnull=True)
            .values("secteur__id", "secteur__code", "secteur__nom")
            .annotate(
                total=Sum("quantite"),
                grands=Coalesce(Sum("quantite", filter=Q(ligne__regime_type="grands")), 0),
                moyens=Coalesce(Sum("quantite", filter=Q(ligne__regime_type="moyens")), 0),
                petits=Coalesce(Sum("quantite", filter=Q(ligne__regime_type="petits")), 0),
            )
            .order_by("-total")
        )
        secteurs = [
            {
                "id": r["secteur__id"], "code": r["secteur__code"], "nom": r["secteur__nom"],
                "total_regimes": int(r["total"] or 0),
                "grands": int(r["grands"] or 0),
                "moyens": int(r["moyens"] or 0),
                "petits": int(r["petits"] or 0),
            }
            for r in secteurs_qs
        ]

        fiches = (
            FicheRecolte.objects.filter(lignes__recolteur=personne, date__year=year, statut="valide", **fiche_extra)
            .prefetch_related(
                Prefetch("lignes",
                         queryset=FicheRecolteLigne.objects.filter(recolteur=personne).prefetch_related("details"))
            )
            .order_by("-date").distinct()[:20]
        )
        last_rows = []
        for f in fiches:
            totals = {"grands": 0, "moyens": 0, "petits": 0}
            prix_total = 0
            salaire_fiche = float(
                FicheRecolteLigne.objects.filter(recolteur=personne, fiche=f)
                .aggregate(t=Coalesce(Sum("salaire_calcule"), Decimal("0")))["t"] or 0
            )
            rates = {k: int(getattr(f, f"bareme_{k}", 0) or 0) for k in totals}
            for line in f.lignes.all():
                qty = sum(int(d.quantite or 0) for d in line.details.all())
                totals[line.regime_type] = totals.get(line.regime_type, 0) + qty
                prix_total += qty * rates.get(line.regime_type, 0)
            total_reg = sum(totals.values())
            last_rows.append({
                "fiche_id": f.id, "date": str(f.date), "statut": f.statut,
                "grands": totals["grands"], "moyens": totals["moyens"],
                "petits": totals["petits"], "total_regimes": total_reg,
                "prix_fcfa": prix_total, "salaire_fcfa": salaire_fiche,
            })

        return Response({
            "recolteur": {
                "id": personne.id,
                "nom": personne.nom,
                "lieu_residence": personne.lieu_residence,
                "numero_telephone": personne.numero_telephone,
                "whatsapp_actif": personne.whatsapp_actif,
                "est_wave": personne.est_wave,
                "date_naissance": str(personne.date_naissance) if personne.date_naissance else None,
            },
            "filtre_superviseur": {
                "actif": cb_user is not None,
                "user_id": cb_user.id if cb_user else None,
                "nom_complet": f"{cb_user.first_name} {cb_user.last_name}".strip() or cb_user.username if cb_user else None,
            },
            "year": year,
            "monthly": {"current": monthly_totals(year), "previous": monthly_totals(prev_year)},
            "yearly": {
                "labels": yearly_labels,
                "data":   [yearly_map.get(y, 0)       for y in yearly_labels],
                "grands": [grands_by_year.get(y, 0)   for y in yearly_labels],
                "moyens": [moyens_by_year.get(y, 0)   for y in yearly_labels],
                "petits": [petits_by_year.get(y, 0)   for y in yearly_labels],
            },
            "stats": {**base, "salaire_total": salaire_total},
            "rang": rang,
            "nb_recolteurs_actifs": nb_recolteurs_actifs,
            "secteurs": secteurs,
            "last_fiches": last_rows,
        })

    @action(detail=False, methods=["get"], url_path="export")
    def export_list(self, request):
        """Export Excel — liste du personnel avec statistiques annuelles."""
        today = timezone.now().date()
        year = int(request.query_params.get("year", today.year))

        wb, ws, _ = _xlsx_response(f"personnel_export_{year}.xlsx")
        headers = ["Téléphone", "Nom", "Lieu résidence", "Wave",
                   "Grands", "Moyens", "Petits", "Total régimes", "Nb fiches", "Dernière récolte"]
        _style_header(ws, headers)

        qs = (
            Personnel.objects.annotate(
                grands=Coalesce(Sum("lignes_recolte__details__quantite",
                                    filter=Q(lignes_recolte__fiche__date__year=year, lignes_recolte__regime_type="grands")), 0),
                moyens=Coalesce(Sum("lignes_recolte__details__quantite",
                                    filter=Q(lignes_recolte__fiche__date__year=year, lignes_recolte__regime_type="moyens")), 0),
                petits=Coalesce(Sum("lignes_recolte__details__quantite",
                                    filter=Q(lignes_recolte__fiche__date__year=year, lignes_recolte__regime_type="petits")), 0),
                total_regimes=Coalesce(Sum("lignes_recolte__details__quantite",
                                           filter=Q(lignes_recolte__fiche__date__year=year)), 0),
                fiches_count=Count("lignes_recolte__fiche", distinct=True,
                                   filter=Q(lignes_recolte__fiche__date__year=year)),
                last_recolte=Max("lignes_recolte__fiche__date", filter=Q(lignes_recolte__fiche__date__year=year)),
            ).order_by("nom")
        )

        totals = [0] * 4
        for row_idx, p in enumerate(qs, start=2):
            g = int(p.grands or 0)
            m = int(p.moyens or 0)
            pe = int(p.petits or 0)
            tot = int(p.total_regimes or 0)
            totals[0] += g; totals[1] += m; totals[2] += pe; totals[3] += tot
            ws.append([
                p.numero_telephone or "", p.nom, p.lieu_residence,
                "Oui" if p.est_wave else "Non",
                g, m, pe, tot,
                int(p.fiches_count or 0),
                str(p.last_recolte) if p.last_recolte else "",
            ])

        # Ligne totaux
        total_row = ws.max_row + 1
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        for i, val in enumerate(totals, start=5):
            cell = ws.cell(row=total_row, column=i, value=val)
            cell.font = Font(bold=True)

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].auto_size = True

        return _send_wb(wb, f"personnel_export_{year}.xlsx")

import io
import json

from django.http import HttpResponse
from django.db.models import Sum, Count, Max, Q, Prefetch
from django.db.models.functions import Coalesce
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response

from recolteurs.models import Personnel
from utils.permissions import IsAdmin, has_droit
from accounts.utils import create_notification
from .models import ActionLog, Client, FicheRecolte, FicheRecolteDetail, FicheRecolteLigne, FicheRecuVente, ParametreBonus
from .serializers import (
    ActionLogSerializer,
    ClientSerializer,
    FicheRecolteSerializer,
    FicheRecolteListSerializer,
    FicheRecuVenteSerializer,
    ParametreBonusSerializer,
)


class ClientViewSet(viewsets.ModelViewSet):
    """
    Lecture : tous les utilisateurs authentifiés.
    Écriture : administrateur ou superviseur avec droit gerer_clients.
    """
    queryset = Client.objects.all().order_by("nom")
    serializer_class = ClientSerializer

    def get_permissions(self):
        if self.action in ("create", "update", "partial_update", "destroy"):
            if has_droit(self.request.user, "gerer_clients"):
                return super().get_permissions()
            return [IsAdmin()]
        return super().get_permissions()

    def create(self, request, *args, **kwargs):
        response = super().create(request, *args, **kwargs)
        if response.status_code in (200, 201):
            snap = {label: str(response.data.get(field) or "") for field, label in _CLIENT_FIELDS.items()}
            _log_action(request.user, "creation_client",
                        detail=f"Client « {response.data.get('nom', '')} » créé.",
                        meta={"snapshot": snap})
        return response

    def partial_update(self, request, *args, **kwargs):
        instance = self.get_object()
        changes = []
        for field, label in _CLIENT_FIELDS.items():
            if field in request.data:
                old_val = getattr(instance, field, None)
                new_val = request.data[field]
                if str(old_val or "") != str(new_val or ""):
                    changes.append({"field": label, "old": str(old_val or ""), "new": str(new_val or "")})
        response = super().partial_update(request, *args, **kwargs)
        if response.status_code == 200:
            _log_action(request.user, "modification_client",
                        detail=f"Client « {instance.nom} » modifié.",
                        meta={"changes": changes})
        return response

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        snap = {label: str(getattr(instance, field, "") or "") for field, label in _CLIENT_FIELDS.items()}
        response = super().destroy(request, *args, **kwargs)
        if response.status_code in (200, 204):
            _log_action(request.user, "suppression_client",
                        detail=f"Client « {instance.nom} » supprimé.",
                        meta={"snapshot": snap})
        return response


def _send_wb(wb, filename: str) -> HttpResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def _style_header(ws, headers, fill_color="1F4E79"):
    fill = PatternFill(fill_type="solid", fgColor=fill_color)
    font = Font(bold=True, color="FFFFFF")
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = font; c.fill = fill; c.alignment = Alignment(horizontal="center")
    ws.auto_filter.ref = ws.dimensions


def _is_admin(user):
    try:
        return user.profile.is_admin
    except AttributeError:
        return False


def _is_superviseur(user):
    """Retourne True uniquement pour le rôle 'superviseur' (pas adjoint)."""
    try:
        return user.profile.role == "superviseur"
    except AttributeError:
        return False


# Champs tracés pour les détails de modification
_FICHE_FIELDS = {
    "date": "Date",
    "superviseur_general": "Superviseur général",
    "bareme_grands": "Barème grands (FCFA)",
    "bareme_moyens": "Barème moyens (FCFA)",
    "bareme_petits": "Barème petits (FCFA)",
    "depense_nourriture": "Dépense nourriture (FCFA)",
    "depense_transport": "Dépense transport (FCFA)",
    "depense_salaire": "Dépense salaire (FCFA)",
    "observations": "Observations",
    "heure_debut": "Heure début",
    "heure_fin": "Heure fin",
    "conditions_meteo": "Conditions météo",
    "nb_palmiers_recoltes": "Palmiers récoltés",
    "surface_recoltee_ha": "Surface récoltée (ha)",
}

_RECU_FIELDS = {
    "date": "Date",
    "client": "Client",
    "pesee_kg": "Pesée (kg)",
    "non_conformes_pct": "Non conformes (%)",
    "montant": "Montant (FCFA)",
    "prix_officiel": "Prix officiel (FCFA/kg)",
    "reference_facture": "Référence facture",
    "mode_paiement": "Mode paiement",
    "vehicule_transport": "Véhicule transport",
}

_CLIENT_FIELDS = {
    "nom": "Nom",
    "telephone": "Téléphone",
    "adresse": "Adresse",
}


def _log_action(acteur, action, fiche=None, recu=None, detail="", meta=None):
    """Enregistre une action dans le journal de traçabilité."""
    superviseur = None
    if fiche and fiche.created_by:
        superviseur = fiche.created_by
    elif recu and recu.fiche and recu.fiche.created_by:
        superviseur = recu.fiche.created_by
    # Quand le superviseur agit sur ses propres données, il est à la fois acteur et superviseur
    if superviseur is None and acteur and _is_superviseur(acteur):
        superviseur = acteur
    if meta:
        stored = {"label": detail}
        stored.update(meta)
        stored_detail = json.dumps(stored, ensure_ascii=False, default=str)
    else:
        stored_detail = detail
    ActionLog.objects.create(
        acteur=acteur,
        superviseur=superviseur,
        action=action,
        fiche=fiche,
        recu=recu,
        detail=stored_detail,
    )


def _notify_statut_recolte(instance, old_statut, new_statut, actor=None):
    if not new_statut or new_statut == old_statut:
        return
    date_str = str(instance.date) if instance.date else f"#{instance.id}"

    if new_statut == "valide":
        actor_is_superviseur = actor and _is_superviseur(actor)
        if actor_is_superviseur:
            # Le superviseur valide → notifier tous les administrateurs
            from django.contrib.auth.models import User as DjangoUser
            actor_name = f"{actor.first_name} {actor.last_name}".strip() or actor.username
            admins = DjangoUser.objects.filter(profile__role="admin", is_active=True)
            for admin in admins:
                create_notification(
                    admin,
                    f"Nouvelle fiche du {date_str} validee par {actor_name}.",
                    "info",
                    "/recoltes",
                )
        elif instance.created_by:
            # L'admin valide → notifier le créateur
            create_notification(
                instance.created_by,
                f"Votre fiche de recolte du {date_str} a ete validee.",
                "success",
                "/recoltes",
            )
    elif new_statut == "brouillon" and old_statut in ("soumis", "valide") and instance.created_by:
        create_notification(
            instance.created_by,
            f"Votre fiche de recolte du {date_str} a ete rejetee.",
            "warning",
            f"/mon-audit?action=rejet&fiche_id={instance.id}",
        )


class FicheRecolteViewSet(viewsets.ModelViewSet):
    serializer_class = FicheRecolteSerializer

    def get_serializer_class(self):
        if self.action == "list":
            return FicheRecolteListSerializer
        return FicheRecolteSerializer

    def get_queryset(self):
        qs = FicheRecolte.objects.select_related(
            "superviseur_general_obj", "created_by", "validated_by"
        )
        if self.action == "list":
            qs = qs.prefetch_related(
                Prefetch("lignes", queryset=FicheRecolteLigne.objects.prefetch_related("details")),
                "recus",
            )
        else:
            qs = qs.prefetch_related("superviseurs_adjoints", "lignes__details", "recus")
        qs = qs.order_by("-id")
        # Admin voit toutes les fiches ; superviseur uniquement les siennes
        if not _is_admin(self.request.user):
            qs = qs.filter(created_by=self.request.user)
        return qs

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    def create(self, request, *args, **kwargs):
        if _is_admin(request.user):
            return Response(
                {"detail": "Les administrateurs ne peuvent pas créer de fiches de récolte. Cette action est réservée aux superviseurs."},
                status=403,
            )
        response = super().create(request, *args, **kwargs)
        if response.status_code in (200, 201):
            fiche_id = response.data.get("id")
            date_str = response.data.get("date", f"#{fiche_id}")
            snap = {
                "Date": str(response.data.get("date", "")),
                "Superviseur général": str(response.data.get("superviseur_general", "") or ""),
                "Barème grands (FCFA)": str(response.data.get("bareme_grands", "")),
                "Barème moyens (FCFA)": str(response.data.get("bareme_moyens", "")),
                "Barème petits (FCFA)": str(response.data.get("bareme_petits", "")),
                "Observations": str(response.data.get("observations", "") or ""),
            }
            _log_action(request.user, "creation_fiche",
                        detail=f"Fiche du {date_str} créée.",
                        meta={"snapshot": snap})
        return response

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        if instance.statut == "valide":
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("Une fiche validée ne peut pas être supprimée.")
        if not _is_admin(request.user) and instance.created_by != request.user:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("Vous ne pouvez supprimer que vos propres fiches.")
        date_str = str(instance.date) if instance.date else f"#{instance.pk}"
        snap = {
            "Date": date_str,
            "Superviseur général": instance.superviseur_general or "",
            "Statut": instance.get_statut_display(),
            "Barème grands (FCFA)": str(instance.bareme_grands),
            "Barème moyens (FCFA)": str(instance.bareme_moyens),
            "Barème petits (FCFA)": str(instance.bareme_petits),
            "Dépense nourriture (FCFA)": str(instance.depense_nourriture),
            "Dépense transport (FCFA)": str(instance.depense_transport),
            "Observations": instance.observations or "",
        }
        _log_action(request.user, "suppression_fiche",
                    detail=f"Fiche du {date_str} supprimée.",
                    meta={"snapshot": snap})
        return super().destroy(request, *args, **kwargs)

    def get_permissions(self):
        if self.action in ("export",):
            return [IsAdmin()]
        return super().get_permissions()

    def _check_statut_transition(self, request, instance=None):
        """Vérifie les règles de transition de statut."""
        new_statut = request.data.get("statut")
        if not new_statut:
            return None
        is_admin = _is_admin(request.user)
        is_sup   = _is_superviseur(request.user)
        can_validate = is_admin or is_sup

        if new_statut == "valide" and not can_validate:
            return Response(
                {"detail": "Seul l'administrateur ou le superviseur peut valider une fiche."},
                status=403,
            )
        # Seul l'admin peut rejeter une fiche (brouillon depuis soumis ou valide)
        if new_statut == "brouillon" and instance and instance.statut in ("soumis", "valide") and not is_admin:
            return Response(
                {"detail": "Seul l'administrateur peut rejeter une fiche de récolte."},
                status=403,
            )
        # Seul l'admin peut modifier une fiche déjà validée
        if instance and instance.statut == "valide" and not is_admin:
            return Response(
                {"detail": "Seul l'administrateur peut modifier une fiche validée."},
                status=403,
            )
        if instance and instance.statut == "valide" and is_admin:
            instance._audit_user  = request.user
            instance._audit_motif = request.data.get("motif", "")
        return None

    def _log_fiche_action(self, request, instance, old_statut, new_statut, data):
        """Détermine et enregistre l'action dans le journal."""
        is_admin = _is_admin(request.user)
        is_sup   = _is_superviseur(request.user)

        # ── Transitions de statut ──────────────────────────────────────────────
        if new_statut and new_statut != old_statut:
            if new_statut == "valide":
                snap = {
                    "Date": str(instance.date),
                    "Superviseur général": instance.superviseur_general or "",
                    "Barème grands (FCFA)": str(instance.bareme_grands),
                    "Barème moyens (FCFA)": str(instance.bareme_moyens),
                    "Barème petits (FCFA)": str(instance.bareme_petits),
                    "Dépense totale (FCFA)": str(instance.depense_total),
                }
                _log_action(request.user, "validation", fiche=instance,
                            detail=f"Fiche du {instance.date} validée.",
                            meta={"snapshot": snap})
            elif new_statut == "soumis" and is_sup:
                snap = {
                    "Date": str(instance.date),
                    "Superviseur général": instance.superviseur_general or "",
                    "Statut précédent": old_statut,
                }
                _log_action(request.user, "soumission_fiche", fiche=instance,
                            detail=f"Fiche du {instance.date} soumise pour validation.",
                            meta={"snapshot": snap})
            elif new_statut == "brouillon" and old_statut in ("soumis", "valide") and is_admin:
                motif = data.get("motif", "").strip()
                detail_rejet = f"Fiche du {instance.date} rejetée."
                if motif:
                    detail_rejet += f" Motif : {motif}"
                _log_action(request.user, "rejet", fiche=instance,
                            detail=detail_rejet,
                            meta={"motif": motif, "date_fiche": str(instance.date)})
            return

        if not is_admin:
            return

        # ── Actions admin uniquement ───────────────────────────────────────────
        bareme_keys = {"bareme_grands", "bareme_moyens", "bareme_petits"}
        changed_bareme = bareme_keys & set(data.keys())
        if changed_bareme:
            details = []
            for k in sorted(changed_bareme):
                old = getattr(instance, k, "?")
                details.append(f"{k} : {old} → {data[k]}")
            _log_action(request.user, "modification_bareme", fiche=instance,
                        detail="Barème modifié. " + " | ".join(details))
            return
        changes = []
        for field, label in _FICHE_FIELDS.items():
            if field in data:
                old_val = getattr(instance, field, None)
                new_val = data[field]
                if str(old_val if old_val is not None else "") != str(new_val if new_val is not None else ""):
                    changes.append({
                        "field": label,
                        "old": str(old_val) if old_val is not None else "",
                        "new": str(new_val) if new_val is not None else "",
                    })
        _log_action(request.user, "modification_fiche", fiche=instance,
                    detail=f"Fiche du {instance.date} modifiée par l'admin.",
                    meta={"changes": changes})

    def partial_update(self, request, *args, **kwargs):
        instance = self.get_object()
        err = self._check_statut_transition(request, instance)
        if err:
            return err
        old_statut = instance.statut
        new_statut = request.data.get("statut")
        self._log_fiche_action(request, instance, old_statut, new_statut, request.data)
        # Appel direct à super().update() avec partial=True pour éviter que
        # DRF's partial_update() rebondisse sur self.update() et double le log.
        kwargs["partial"] = True
        response = super().update(request, *args, **kwargs)
        if new_statut == "valide" and old_statut != "valide":
            instance.refresh_from_db()
            instance.validated_by = request.user
            instance.validated_at = timezone.now()
            instance.save(update_fields=["validated_by", "validated_at"])
        elif new_statut == "brouillon" and old_statut == "valide":
            instance.refresh_from_db()
            instance.validated_by = None
            instance.validated_at = None
            instance.save(update_fields=["validated_by", "validated_at"])
        _notify_statut_recolte(instance, old_statut, new_statut, actor=request.user)
        return response

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        err = self._check_statut_transition(request, instance)
        if err:
            return err
        old_statut = instance.statut
        new_statut = request.data.get("statut")
        self._log_fiche_action(request, instance, old_statut, new_statut, request.data)
        response = super().update(request, *args, **kwargs)
        if new_statut == "valide" and old_statut != "valide":
            instance.refresh_from_db()
            instance.validated_by = request.user
            instance.validated_at = timezone.now()
            instance.save(update_fields=["validated_by", "validated_at"])
        _notify_statut_recolte(instance, old_statut, new_statut, actor=request.user)
        return response

    @action(detail=False, methods=["get"], url_path="analytics")
    def analytics(self, request):
        today = timezone.now().date()
        year = int(request.query_params.get("year", today.year))
        prev_year = year - 1
        prev2_year = year - 2
        is_admin = _is_admin(request.user)

        # Filtre de base : superviseur ne voit que ses propres fiches
        user_filter = Q() if is_admin else Q(ligne__fiche__created_by=request.user)
        personnel_user_filter = Q() if is_admin else Q(lignes_recolte__fiche__created_by=request.user)

        LABELS_MOIS = ["Jan", "Fev", "Mar", "Avr", "Mai", "Juin", "Juil", "Aout", "Sept", "Oct", "Nov", "Dec"]

        def monthly_totals(target_year):
            qs = (
                FicheRecolteDetail.objects.filter(user_filter, ligne__fiche__date__year=target_year)
                .values("ligne__fiche__date__month", "ligne__regime_type")
                .annotate(total=Sum("quantite"))
            )
            by_month = {}
            for r in qs:
                m = r["ligne__fiche__date__month"]
                t = r["ligne__regime_type"]
                by_month.setdefault(m, {})[t] = int(r["total"] or 0)
            return {
                "labels": LABELS_MOIS,
                "grands": [by_month.get(m, {}).get("grands", 0) for m in range(1, 13)],
                "moyens": [by_month.get(m, {}).get("moyens", 0) for m in range(1, 13)],
                "petits": [by_month.get(m, {}).get("petits", 0) for m in range(1, 13)],
            }

        start_year = year - 4
        yearly_qs = (
            FicheRecolteDetail.objects.filter(user_filter, ligne__fiche__date__year__gte=start_year)
            .values("ligne__fiche__date__year", "ligne__regime_type")
            .annotate(total=Sum("quantite"))
            .order_by("ligne__fiche__date__year")
        )
        yearly_map = {}
        for r in yearly_qs:
            y_ = r["ligne__fiche__date__year"]
            t = r["ligne__regime_type"]
            yearly_map.setdefault(y_, {})[t] = int(r["total"] or 0)
        yearly_labels = list(range(start_year, year + 1))
        yearly = {
            "labels": yearly_labels,
            "grands": [yearly_map.get(y_, {}).get("grands", 0) for y_ in yearly_labels],
            "moyens": [yearly_map.get(y_, {}).get("moyens", 0) for y_ in yearly_labels],
            "petits": [yearly_map.get(y_, {}).get("petits", 0) for y_ in yearly_labels],
        }

        year_q = Q(lignes_recolte__fiche__date__year=year) & personnel_user_filter
        personnel = (
            Personnel.objects.annotate(
                grands=Coalesce(Sum("lignes_recolte__details__quantite",
                                    filter=year_q & Q(lignes_recolte__regime_type="grands")), 0),
                moyens=Coalesce(Sum("lignes_recolte__details__quantite",
                                    filter=year_q & Q(lignes_recolte__regime_type="moyens")), 0),
                petits=Coalesce(Sum("lignes_recolte__details__quantite",
                                    filter=year_q & Q(lignes_recolte__regime_type="petits")), 0),
                total_regimes=Coalesce(Sum("lignes_recolte__details__quantite",
                                           filter=year_q), 0),
                fiches_count=Count("lignes_recolte__fiche", distinct=True, filter=year_q),
                last_recolte=Max("lignes_recolte__fiche__date", filter=year_q),
            )
            .values("id", "numero_telephone", "nom", "lieu_residence",
                    "grands", "moyens", "petits", "total_regimes", "fiches_count", "last_recolte")
            .order_by("-total_regimes", "nom")
        )
        # Pour un superviseur : n'afficher que les récolteurs qui ont participé à ses fiches
        if not is_admin:
            personnel = personnel.filter(
                lignes_recolte__fiche__created_by=request.user,
                lignes_recolte__fiche__date__year=year,
            ).distinct()

        return Response({
            "year": year,
            "monthly": {
                "current":  monthly_totals(year),
                "previous": monthly_totals(prev_year),
                "prev2":    monthly_totals(prev2_year),
            },
            "yearly": yearly,
            "recolteurs": list(personnel),
        })

    @action(detail=False, methods=["get"], url_path="by_date")
    def by_date(self, request):
        date_str = request.query_params.get("date")
        if not date_str:
            return Response({"detail": "Paramètre 'date' requis."}, status=400)

        qs = FicheRecolte.objects.filter(date=date_str)
        if not _is_admin(request.user):
            qs = qs.filter(created_by=request.user)
        fiches = qs.prefetch_related(
            "lignes__recolteur", "lignes__details__secteur", "superviseurs_adjoints"
        ).order_by("id")

        result = []
        for fiche in fiches:
            lignes_data = []
            for ligne in fiche.lignes.all():
                det_list = [
                    {
                        "secteur_code": d.secteur_code or (d.secteur.code if d.secteur else ""),
                        "quantite": int(d.quantite),
                    }
                    for d in ligne.details.all()
                ]
                lignes_data.append({
                    "recolteur_nom": (ligne.recolteur.nom if ligne.recolteur else None) or ligne.recolteur_nom or "—",
                    "recolteur_telephone": ligne.recolteur.numero_telephone if ligne.recolteur else "",
                    "regime_type": ligne.regime_type,
                    "details": det_list,
                    "total": sum(d["quantite"] for d in det_list),
                })

            grands = sum(l["total"] for l in lignes_data if l["regime_type"] == "grands")
            moyens = sum(l["total"] for l in lignes_data if l["regime_type"] == "moyens")
            petits = sum(l["total"] for l in lignes_data if l["regime_type"] == "petits")

            adjoints = [
                {"nom": s.nom, "secteur_ou_recolteur": s.secteur_ou_recolteur}
                for s in fiche.superviseurs_adjoints.all()
            ]

            result.append({
                "id": fiche.id,
                "date": str(fiche.date),
                "statut": fiche.statut,
                "superviseur_general": fiche.superviseur_general or "—",
                "superviseurs_adjoints": adjoints,
                "total_regimes": grands + moyens + petits,
                "grands": grands,
                "moyens": moyens,
                "petits": petits,
                "lignes": lignes_data,
            })

        return Response(result)

    @action(detail=False, methods=["get"], url_path="export")
    def export(self, request):
        """Export Excel des fiches de récolte."""
        year = request.query_params.get("year")
        secteur_id = request.query_params.get("secteur")
        today = timezone.now().date()
        year = int(year) if year else today.year

        wb = Workbook()
        ws = wb.active
        ws.title = f"Recoltes {year}"
        headers = ["Date", "Tél. récolteur", "Nom récolteur", "Type régime",
                   "Code secteur", "Quantité", "Fiche ID"]
        _style_header(ws, headers)

        details_qs = FicheRecolteDetail.objects.select_related(
            "ligne__fiche", "ligne__recolteur", "secteur"
        ).filter(ligne__fiche__date__year=year)
        if secteur_id:
            details_qs = details_qs.filter(secteur_id=secteur_id)

        details = (
            details_qs.values(
                "ligne__fiche__date", "ligne__recolteur__numero_telephone",
                "ligne__recolteur__nom", "ligne__recolteur_nom",
                "ligne__regime_type", "secteur__code", "secteur_code",
                "quantite", "ligne__fiche__id",
            ).order_by("ligne__fiche__date")
        )

        total_qty = 0
        for row in details:
            nom = row["ligne__recolteur__nom"] or row["ligne__recolteur_nom"] or ""
            scode = row["secteur__code"] or row["secteur_code"] or ""
            qty = int(row["quantite"] or 0)
            total_qty += qty
            ws.append([str(row["ligne__fiche__date"]), row["ligne__recolteur__numero_telephone"] or "",
                        nom, row["ligne__regime_type"], scode, qty, row["ligne__fiche__id"]])

        total_row = ws.max_row + 1
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=6, value=total_qty).font = Font(bold=True)

        return _send_wb(wb, f"recoltes_export_{year}.xlsx")


class FicheRecuVenteViewSet(viewsets.ModelViewSet):
    """
    CRUD standalone sur les reçus de vente.
    - Superviseur : voit et gère uniquement les reçus de ses propres fiches.
    - Admin : voit tout, peut modifier prix_officiel, valider/modifier les reçus validés.
    """
    serializer_class = FicheRecuVenteSerializer

    def get_queryset(self):
        qs = FicheRecuVente.objects.select_related(
            "fiche", "fiche__superviseur_general_obj", "client_obj", "validated_by"
        ).order_by("-date", "-id")
        if not _is_admin(self.request.user) and not has_droit(self.request.user, "gerer_recus_vente"):
            qs = qs.filter(fiche__created_by=self.request.user)

        fiche_id = self.request.query_params.get("fiche")
        if fiche_id:
            qs = qs.filter(fiche_id=fiche_id)
        year = self.request.query_params.get("year")
        if year:
            qs = qs.filter(date__year=year)

        # Filtres de recherche
        search = self.request.query_params.get("search", "").strip()
        if search:
            qs = qs.filter(
                Q(client__icontains=search)
                | Q(client_obj__nom__icontains=search)
                | Q(reference_facture__icontains=search)
            )
        statut = self.request.query_params.get("statut")
        if statut:
            qs = qs.filter(statut=statut)
        return qs

    def create(self, request, *args, **kwargs):
        from rest_framework.exceptions import PermissionDenied
        if not _is_admin(request.user) and not has_droit(request.user, "gerer_recus_vente"):
            raise PermissionDenied("Vous n'avez pas le droit de créer des reçus de vente.")
        response = super().create(request, *args, **kwargs)
        if response.status_code in (200, 201):
            try:
                instance = FicheRecuVente.objects.select_related("fiche").get(pk=response.data["id"])
                snap = {label: str(getattr(instance, field, "") or "") for field, label in _RECU_FIELDS.items()}
                fiche_ref = instance.fiche
            except Exception:
                snap = {}
                fiche_ref = None
            _log_action(request.user, "creation_recu",
                        fiche=fiche_ref,
                        detail=f"Reçu #{response.data.get('id', '')} créé (fiche du {response.data.get('fiche_date', '?')}).",
                        meta={"snapshot": snap})
        return response

    def _check_ownership(self, serializer):
        from rest_framework.exceptions import PermissionDenied
        fiche = serializer.validated_data.get("fiche")
        if fiche and not _is_admin(self.request.user):
            if fiche.created_by != self.request.user:
                raise PermissionDenied("Vous ne pouvez créer un reçu que pour vos propres fiches.")

    def perform_create(self, serializer):
        self._check_ownership(serializer)
        extra = {}
        # Auto-fill prix_officiel from global parameter if not provided
        if serializer.validated_data.get("prix_officiel") is None:
            param_prix = ParametreBonus.get_instance().prix_kg_officiel
            if param_prix is not None:
                extra["prix_officiel"] = param_prix
        serializer.save(**extra)

    def _check_recu_editable(self, request, instance):
        """Un reçu validé est figé — personne ne peut le modifier (seule la suppression/annulation reste possible)."""
        if instance.statut == "valide":
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("Ce reçu est validé et ne peut plus être modifié. Supprimez-le pour le recréer si nécessaire.")

    def partial_update(self, request, *args, **kwargs):
        instance = self.get_object()
        self._check_recu_editable(request, instance)
        if not _is_admin(request.user) and "prix_officiel" in request.data:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("Seul l'administrateur peut modifier le prix officiel.")
        # Logging admin
        if _is_admin(request.user):
            if "prix_officiel" in request.data:
                old_val = instance.prix_officiel or "—"
                new_val = request.data["prix_officiel"]
                _log_action(request.user, "prix_officiel", fiche=instance.fiche, recu=instance,
                            detail=f"Prix officiel : {old_val} → {new_val} FCFA/kg (reçu #{instance.id}, fiche du {instance.fiche.date if instance.fiche else '?'})")
            else:
                changes = []
                for field, label in _RECU_FIELDS.items():
                    if field in request.data:
                        old_val = getattr(instance, field, None)
                        new_val = request.data[field]
                        if str(old_val if old_val is not None else "") != str(new_val if new_val is not None else ""):
                            changes.append({
                                "field": label,
                                "old": str(old_val) if old_val is not None else "",
                                "new": str(new_val) if new_val is not None else "",
                            })
                _log_action(request.user, "modification_recu", fiche=instance.fiche, recu=instance,
                            detail=f"Reçu #{instance.id} modifié (fiche du {instance.fiche.date if instance.fiche else '?'}).",
                            meta={"changes": changes})
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        if instance.statut == "valide" and not _is_admin(request.user):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("Ce reçu est validé et ne peut être supprimé que par l'administrateur.")
        snap = {label: str(getattr(instance, field, "") or "") for field, label in _RECU_FIELDS.items()}
        _log_action(request.user, "suppression_recu", fiche=instance.fiche, recu=instance,
                    detail=f"Reçu #{instance.id} du {instance.date} supprimé (fiche du {instance.fiche.date if instance.fiche else '?'}).",
                    meta={"snapshot": snap})
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=["post"], url_path="valider")
    def valider(self, request, pk=None):
        """Valide un reçu de vente — le superviseur créateur ou un admin."""
        instance = self.get_object()  # queryset limite déjà l'accès aux propres fiches
        can_validate = (
            _is_admin(request.user)
            or has_droit(request.user, "gerer_recus_vente")
            or (instance.fiche and instance.fiche.created_by == request.user)
        )
        if not can_validate:
            return Response({"detail": "Vous n'avez pas le droit de valider ce reçu."}, status=403)
        if instance.statut == "valide":
            return Response({"detail": "Ce reçu est déjà validé."}, status=400)
        instance.statut = "valide"
        instance.validated_by = request.user
        instance.validated_at = timezone.now()
        instance.save(update_fields=["statut", "validated_by", "validated_at"])
        snap = {label: str(getattr(instance, field, "") or "") for field, label in _RECU_FIELDS.items()}
        _log_action(
            request.user, "validation_recu",
            fiche=instance.fiche, recu=instance,
            detail=f"Reçu #{instance.id} du {instance.date} validé (fiche du {instance.fiche.date if instance.fiche else '?'}).",
            meta={"snapshot": snap},
        )
        if not _is_admin(request.user):
            from django.contrib.auth.models import User as DjangoUser
            actor_name = f"{request.user.first_name} {request.user.last_name}".strip() or request.user.username
            admins = DjangoUser.objects.filter(profile__role="admin", is_active=True)
            for admin in admins:
                create_notification(
                    admin,
                    f"Le reçu #{instance.id} du {instance.date} a été validé par {actor_name}.",
                    "info",
                    "/recoltes?tab=ventes",
                )
        return Response(self.get_serializer(instance).data)

    @action(detail=True, methods=["post"], url_path="rejeter")
    def rejeter(self, request, pk=None):
        """Rejette un reçu validé — le repasse en brouillon (admin uniquement)."""
        if not _is_admin(request.user):
            return Response({"detail": "Seul l'administrateur peut rejeter un reçu."}, status=403)
        instance = self.get_object()
        if instance.statut != "valide":
            return Response({"detail": "Seul un reçu validé peut être rejeté."}, status=400)

        motif = (request.data.get("motif") or "").strip()
        instance.statut = "valide"   # temporaire pour l'audit snapshot
        snap = {label: str(getattr(instance, field, "") or "") for field, label in _RECU_FIELDS.items()}

        instance.statut = "brouillon"
        instance.validated_by = None
        instance.validated_at = None
        instance.save(update_fields=["statut", "validated_by", "validated_at"])

        detail_rejet = f"Reçu #{instance.id} du {instance.date} rejeté (fiche du {instance.fiche.date if instance.fiche else '?'})."
        if motif:
            detail_rejet += f" Motif : {motif}"
        _log_action(request.user, "rejet_recu", fiche=instance.fiche, recu=instance,
                    detail=detail_rejet,
                    meta={"snapshot": snap, "motif": motif})

        if instance.fiche and instance.fiche.created_by:
            create_notification(
                instance.fiche.created_by,
                f"Votre reçu #{instance.id} du {instance.date} a été rejeté.",
                "warning",
                f"/mon-audit?action=rejet_recu&fiche_id={instance.fiche.id}",
            )

        return Response(self.get_serializer(instance).data)


class ParametreBonusViewSet(viewsets.ViewSet):
    """
    Singleton admin : GET pour lire, PATCH pour modifier.
    """
    def list(self, request):
        obj = ParametreBonus.get_instance()
        return Response(ParametreBonusSerializer(obj).data)

    def partial_update(self, request, pk=None):
        if not _is_admin(request.user):
            return Response({"detail": "Réservé à l'administrateur."}, status=403)
        obj = ParametreBonus.get_instance()
        _BONUS_FIELDS = {
            "bareme_grands_defaut":  "Barème grands défaut (FCFA)",
            "bareme_moyens_defaut":  "Barème moyens défaut (FCFA)",
            "bareme_petits_defaut":  "Barème petits défaut (FCFA)",
            "seuil_non_conformes":   "Seuil non conformes (%)",
            "montant_bonus":         "Montant bonus (FCFA)",
        }
        changes = []
        for field, label in _BONUS_FIELDS.items():
            if field in request.data:
                old_val = getattr(obj, field, None)
                new_val = request.data[field]
                if str(old_val or "") != str(new_val or ""):
                    changes.append({"field": label, "old": str(old_val or ""), "new": str(new_val or "")})
        serializer = ParametreBonusSerializer(obj, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        _log_action(request.user, "modification_bareme",
                    detail="Paramètres bonus modifiés.",
                    meta={"changes": changes})
        return Response(serializer.data)


class ActionLogViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Journal d'audit.
    - Admin : voit tous les logs, peut filtrer par superviseur, peut annuler.
    - Superviseur : voit uniquement les logs où il est concerné.
    """
    serializer_class = ActionLogSerializer

    @action(detail=True, methods=["post"], url_path="annuler", permission_classes=[IsAdmin])
    def annuler(self, request, pk=None):
        import json
        import datetime
        from decimal import Decimal, InvalidOperation
        from django.apps import apps
        from accounts.models import Notification

        try:
            log = ActionLog.objects.select_related("acteur", "superviseur", "fiche").get(pk=pk)
        except ActionLog.DoesNotExist:
            return Response({"detail": "Action introuvable."}, status=404)

        if log.annule:
            return Response({"detail": "Cette action a déjà été annulée."}, status=400)

        REVERSIBLE_REF = {
            "modification_secteur":  ("secteurs",  "Secteur"),
            "modification_agent":    ("agents",    "AgentTerrain"),
            "modification_materiel": ("materiels", "MaterielEquipement"),
            "modification_recolteur":("recolteurs","Personnel"),
        }
        if log.action not in REVERSIBLE_REF and log.action not in ("soumission_fiche", "suppression_recolteur"):
            return Response({"detail": "Ce type d'action ne peut pas être annulé."}, status=400)

        raison = (request.data.get("raison") or "").strip()

        try:
            parsed = json.loads(log.detail) if log.detail else {}
        except Exception:
            parsed = {}

        revert_label = parsed.get("label", log.get_action_display())

        # ── Revert suppression récolteur ─────────────────────────────
        if log.action == "suppression_recolteur":
            before_raw = parsed.get("before_raw")
            if not before_raw or "id" not in before_raw:
                return Response(
                    {"detail": "Données de restauration manquantes — cette entrée a été créée avant l'activation de l'annulation."},
                    status=400,
                )
            from recolteurs.models import Personnel
            from recoltes.models import FicheRecolteLigne

            original_id = before_raw["id"]
            nom = before_raw.get("nom") or ""

            if Personnel.objects.filter(pk=original_id).exists():
                return Response({"detail": "Un récolteur avec cet identifiant existe déjà en base."}, status=400)

            date_naissance = None
            if before_raw.get("date_naissance"):
                try:
                    date_naissance = datetime.date.fromisoformat(str(before_raw["date_naissance"]))
                except (ValueError, TypeError):
                    pass

            try:
                Personnel.objects.create(
                    id=original_id,
                    nom=nom,
                    lieu_residence=before_raw.get("lieu_residence") or "",
                    numero_telephone=before_raw.get("numero_telephone") or "",
                    whatsapp_actif=bool(before_raw.get("whatsapp_actif", False)),
                    est_wave=bool(before_raw.get("est_wave", False)),
                    date_naissance=date_naissance,
                )
            except Exception as e:
                return Response({"detail": f"Impossible de restaurer le récolteur : {e}"}, status=400)

            # Relier exactement les lignes qui appartenaient à ce récolteur (par IDs sauvegardés)
            ligne_ids = before_raw.get("ligne_ids") or []
            if ligne_ids:
                FicheRecolteLigne.objects.filter(
                    id__in=ligne_ids,
                    recolteur__isnull=True,
                ).update(recolteur_id=original_id)
            else:
                # Fallback pour les entrées créées avant ce correctif
                FicheRecolteLigne.objects.filter(
                    recolteur__isnull=True,
                    recolteur_nom=nom,
                ).update(recolteur_id=original_id)

        # ── Revert soumission fiche ──────────────────────────────────
        elif log.action == "soumission_fiche":
            fiche = log.fiche
            if not fiche:
                return Response({"detail": "Fiche introuvable."}, status=404)
            if fiche.statut == "brouillon":
                return Response({"detail": "La fiche est déjà en brouillon."}, status=400)
            fiche.statut = "brouillon"
            fiche.validated_by = None
            fiche.validated_at = None
            fiche.save(update_fields=["statut", "validated_by", "validated_at"])

        # ── Revert modification référentiel ──────────────────────────
        else:
            app_label, model_name = REVERSIBLE_REF[log.action]
            object_id  = parsed.get("object_id")
            before_raw = parsed.get("before_raw", {})

            if not object_id:
                return Response(
                    {"detail": "ID manquant — cette entrée a été créée avant l'activation de l'annulation."},
                    status=400,
                )
            if not before_raw:
                return Response(
                    {"detail": "Données de restauration manquantes — entrée créée avant l'activation de l'annulation."},
                    status=400,
                )

            Model = apps.get_model(app_label, model_name)
            try:
                obj = Model.objects.get(pk=object_id)
            except Model.DoesNotExist:
                return Response({"detail": "L'objet n'existe plus dans la base de données."}, status=404)

            def _apply(obj, field_name, val):
                if not hasattr(obj, field_name):
                    return
                try:
                    fo = obj._meta.get_field(field_name)
                    ft = fo.get_internal_type()
                    nullable = getattr(fo, "null", False)
                    if val is None:
                        setattr(obj, field_name, None if nullable else "")
                    elif ft == "DecimalField":
                        setattr(obj, field_name, Decimal(str(val)))
                    elif ft in ("IntegerField", "PositiveIntegerField", "BigIntegerField", "SmallIntegerField"):
                        setattr(obj, field_name, int(val))
                    elif ft == "BooleanField":
                        setattr(obj, field_name, val if isinstance(val, bool) else str(val).lower() in ("true", "1"))
                    elif ft == "DateField":
                        setattr(obj, field_name, datetime.date.fromisoformat(str(val)) if val else None)
                    elif ft in ("ForeignKey",):
                        setattr(obj, field_name, int(val) if val else None)
                    else:
                        setattr(obj, field_name, str(val) if val is not None else ("" if not nullable else None))
                except (ValueError, InvalidOperation):
                    pass

            for field_name, val in before_raw.items():
                _apply(obj, field_name, val)
            obj.save()

        # ── Marquer log comme annulé ─────────────────────────────────
        log.annule = True
        log.save(update_fields=["annule"])

        # ── Log de l'annulation ──────────────────────────────────────
        ActionLog.objects.create(
            acteur=request.user,
            superviseur=log.acteur,
            action="annulation_action",
            fiche=log.fiche,
            detail=json.dumps({
                "label": f"Annulation : {revert_label}",
                "reverts_id": log.id,
                "raison": raison,
            }, ensure_ascii=False),
        )

        # ── Notification au superviseur ──────────────────────────────
        target = log.acteur
        if target:
            msg = f"L'administrateur a annulé une de vos actions : {revert_label}."
            if raison:
                msg += f" Raison : {raison}"
            Notification.objects.create(
                user=target,
                message=msg,
                type="warning",
                lien="/mon-audit",
            )

        return Response({"ok": True})

    def get_queryset(self):
        is_admin = _is_admin(self.request.user)
        qs = ActionLog.objects.select_related(
            "acteur", "superviseur", "fiche", "recu"
        )

        if not is_admin:
            # Superviseur voit : actions admin sur ses fiches + ses propres actions
            qs = qs.filter(
                Q(superviseur=self.request.user) | Q(acteur=self.request.user)
            )
        else:
            superviseur_id = self.request.query_params.get("superviseur")
            if superviseur_id:
                qs = qs.filter(superviseur_id=superviseur_id)

        action = self.request.query_params.get("action")
        date_debut = self.request.query_params.get("date_debut")
        date_fin   = self.request.query_params.get("date_fin")
        fiche_id   = self.request.query_params.get("fiche_id")

        if action:
            qs = qs.filter(action=action)
        if date_debut:
            qs = qs.filter(timestamp__date__gte=date_debut)
        if date_fin:
            qs = qs.filter(timestamp__date__lte=date_fin)
        if fiche_id:
            qs = qs.filter(fiche_id=fiche_id)
        return qs

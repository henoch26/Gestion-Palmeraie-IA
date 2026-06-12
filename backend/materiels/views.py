import io

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from rest_framework import viewsets
from rest_framework.decorators import action

from utils.permissions import IsAdmin
from utils.audit import log_action, snapshot, diff_fields, build_revert_meta
from .models import MaterielEquipement, MaterielUtiliseTravaux
from .serializers import MaterielEquipementSerializer, MaterielUtiliseTravauxSerializer

_MATERIEL_LABELS = {
    "numero":                    "N°",
    "designation":               "Désignation",
    "quantite":                  "Quantité",
    "categorie":                 "Catégorie",
    "etat_physique":             "État physique",
    "statut_utilisation":        "Statut utilisation",
    "fournisseur":               "Fournisseur",
    "localisation":              "Localisation",
    "responsable":               "Responsable",
    "date_acquisition":          "Date acquisition",
    "valeur_achat":              "Valeur achat",
    "date_derniere_maintenance": "Dernière maintenance",
    "date_prochaine_maintenance":"Prochaine maintenance",
}


def _send_wb(wb, filename: str) -> HttpResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def _style_header(ws, headers, fill_color="1F4E79"):
    fill = PatternFill(fill_type="solid", fgColor=fill_color)
    font = Font(bold=True, color="FFFFFF")
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = font; c.fill = fill; c.alignment = Alignment(horizontal="center")
    ws.auto_filter.ref = ws.dimensions


class MaterielEquipementViewSet(viewsets.ModelViewSet):
    queryset = MaterielEquipement.objects.all().order_by("numero")
    serializer_class = MaterielEquipementSerializer

    def get_permissions(self):
        if self.action in ("export",):
            return [IsAdmin()]
        return super().get_permissions()

    def create(self, request, *args, **kwargs):
        response = super().create(request, *args, **kwargs)
        if response.status_code in (200, 201):
            try:
                inst = MaterielEquipement.objects.get(pk=response.data["id"])
                snap = snapshot(inst, _MATERIEL_LABELS)
            except Exception:
                snap = {}
            designation = response.data.get("designation", "") or f"#{response.data.get('id', '')}"
            log_action(request.user, "creation_materiel",
                       detail=f"Matériel « {designation} » créé.",
                       meta={"snapshot": snap})
        return response

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        pk = instance.pk
        before = snapshot(instance, _MATERIEL_LABELS)
        before_raw = build_revert_meta(instance, _MATERIEL_LABELS)
        response = super().update(request, *args, **kwargs)
        if response.status_code == 200:
            fresh = MaterielEquipement.objects.get(pk=pk)
            after = snapshot(fresh, _MATERIEL_LABELS)
            changes = diff_fields(before, after)
            log_action(request.user, "modification_materiel",
                       detail=f"Matériel « {fresh.designation} » modifié.",
                       meta={"changes": changes, "object_id": pk, "object_type": "materiel", "before_raw": before_raw})
        return response

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        snap = snapshot(instance, _MATERIEL_LABELS)
        log_action(request.user, "suppression_materiel",
                   detail=f"Matériel « {instance.designation} » supprimé.",
                   meta={"snapshot": snap})
        return super().destroy(request, *args, **kwargs)

    @action(detail=False, methods=["get"], url_path="export")
    def export(self, request):
        """Export Excel du stock de matériels."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Matériels"
        headers = ["N°", "Désignation", "Quantité", "Catégorie", "État physique",
                   "Statut utilisation", "Fournisseur", "Date acquisition",
                   "Valeur achat", "Localisation", "Responsable",
                   "Dernière maintenance", "Prochaine maintenance"]
        _style_header(ws, headers)

        total_val = 0
        for m in MaterielEquipement.objects.all().order_by("numero"):
            val = float(m.valeur_achat or 0)
            total_val += val
            ws.append([
                m.numero, m.designation or "", m.quantite,
                m.get_categorie_display() if m.categorie else "",
                m.etat_physique or "", m.statut_utilisation or "",
                m.fournisseur or "",
                str(m.date_acquisition) if m.date_acquisition else "",
                val if val else "",
                m.localisation or "", m.responsable or "",
                str(m.date_derniere_maintenance) if m.date_derniere_maintenance else "",
                str(m.date_prochaine_maintenance) if m.date_prochaine_maintenance else "",
            ])

        tr = ws.max_row + 1
        ws.cell(tr, 1, "TOTAL").font = Font(bold=True)
        ws.cell(tr, 9, round(total_val, 2)).font = Font(bold=True)

        return _send_wb(wb, f"materiels_export_{timezone.now().year}.xlsx")


class MaterielUtiliseTravauxViewSet(viewsets.ModelViewSet):
    queryset = MaterielUtiliseTravaux.objects.select_related("materiel", "fiche_travaux").order_by("-id")
    serializer_class = MaterielUtiliseTravauxSerializer

    def create(self, request, *args, **kwargs):
        response = super().create(request, *args, **kwargs)
        if response.status_code in (200, 201):
            try:
                inst = MaterielUtiliseTravaux.objects.select_related("materiel", "fiche_travaux").get(pk=response.data["id"])
                ref = str(inst.fiche_travaux.periode_travaux) if inst.fiche_travaux and inst.fiche_travaux.periode_travaux else f"fiche #{inst.fiche_travaux_id}"
                snap = {
                    "Matériel": str(inst.materiel) if inst.materiel else "",
                    "Quantité utilisée": str(inst.quantite_utilisee),
                    "Fiche travaux": ref,
                }
            except Exception:
                snap = {}
                ref = ""
            log_action(request.user, "creation_materiel_travaux",
                       detail=f"Matériel ajouté dans la fiche travaux ({ref}).",
                       meta={"snapshot": snap})
        return response

    def partial_update(self, request, *args, **kwargs):
        instance = self.get_object()
        old_qty = str(instance.quantite_utilisee)
        response = super().partial_update(request, *args, **kwargs)
        if response.status_code == 200:
            instance.refresh_from_db()
            new_qty = str(instance.quantite_utilisee)
            ref = str(instance.fiche_travaux.periode_travaux) if instance.fiche_travaux and instance.fiche_travaux.periode_travaux else f"fiche #{instance.fiche_travaux_id}"
            changes = []
            if old_qty != new_qty:
                changes.append({"field": "Quantité utilisée", "old": old_qty, "new": new_qty})
            log_action(request.user, "modification_materiel_travaux",
                       detail=f"Matériel « {instance.materiel} » modifié dans la fiche travaux ({ref}).",
                       meta={"changes": changes})
        return response

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        ref = str(instance.fiche_travaux.periode_travaux) if instance.fiche_travaux and instance.fiche_travaux.periode_travaux else f"fiche #{instance.fiche_travaux_id}"
        snap = {
            "Matériel": str(instance.materiel) if instance.materiel else "",
            "Quantité utilisée": str(instance.quantite_utilisee),
            "Fiche travaux": ref,
        }
        log_action(request.user, "suppression_materiel_travaux",
                   detail=f"Matériel « {instance.materiel} » retiré de la fiche travaux ({ref}).",
                   meta={"snapshot": snap})
        return super().destroy(request, *args, **kwargs)

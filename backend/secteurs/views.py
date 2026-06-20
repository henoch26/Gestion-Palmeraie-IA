"""
secteurs/views.py — API REST pour les secteurs de la palmeraie.

Expose SecteurViewSet (ModelViewSet) :
  GET/POST        /api/secteurs/              → liste et creation
  GET/PATCH/DELETE /api/secteurs/:id/         → detail, modification, suppression
  GET             /api/secteurs/export/       → export Excel de tous les secteurs actifs
  GET             /api/secteurs/:id/analytics/→ indicateurs de production du secteur
  GET             /api/secteurs/:id/export/   → export Excel des recoltes du secteur

Le queryset annote chaque secteur avec :
  production_total — total des regimes recoltes (fiches validees uniquement)
  last_recolte     — date de la derniere recolte validee

Toute creation/modification/suppression est tracee dans ActionLog.
"""
import io

from django.http import HttpResponse
from django.db.models import Q, Sum, Max
from django.db.models.functions import Coalesce
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response

from .models import Secteur
from .serializers import SecteurSerializer
from recoltes.models import FicheRecolte, FicheRecolteDetail
from recolteurs.models import Personnel
from utils.audit import log_action, snapshot, diff_fields, build_revert_meta

_SECTEUR_LABELS = {
    "code":                "Code",
    "nom":                 "Nom",
    "superficie_ha":       "Superficie (ha)",
    "situation_relief":    "Relief",
    "type_sol":            "Type de sol",
    "age_moyen_plants":    "Âge moyen plants (ans)",
    "nb_palmiers":         "Nb palmiers",
    "rendement_cible_t_ha":"Rendement cible (t/ha)",
    "coordonnees_GPS":     "Coordonnées GPS",
    "statut":              "Statut",
}


def _send_wb(wb, filename: str) -> HttpResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def _style_header(ws, headers, fill_color="1F4E79"):
    fill = PatternFill(fill_type="solid", fgColor=fill_color)
    font = Font(bold=True, color="FFFFFF")
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
    ws.auto_filter.ref = ws.dimensions


class SecteurViewSet(viewsets.ModelViewSet):
    serializer_class = SecteurSerializer

    def create(self, request, *args, **kwargs):
        response = super().create(request, *args, **kwargs)
        if response.status_code in (200, 201):
            try:
                inst = Secteur.objects.get(pk=response.data["id"])
                snap = snapshot(inst, _SECTEUR_LABELS)
            except Exception:
                snap = {}
            nom = f"{response.data.get('code', '')} — {response.data.get('nom', '')}".strip(" —")
            log_action(request.user, "creation_secteur",
                       detail=f"Secteur « {nom} » créé.",
                       meta={"snapshot": snap})
        return response

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        pk = instance.pk
        before = snapshot(instance, _SECTEUR_LABELS)
        before_raw = build_revert_meta(instance, _SECTEUR_LABELS)
        response = super().update(request, *args, **kwargs)
        if response.status_code == 200:
            fresh = Secteur.objects.get(pk=pk)
            after = snapshot(fresh, _SECTEUR_LABELS)
            changes = diff_fields(before, after)
            nom = f"{fresh.code} — {fresh.nom}"
            log_action(request.user, "modification_secteur",
                       detail=f"Secteur « {nom} » modifié.",
                       meta={"changes": changes, "object_id": pk, "object_type": "secteur", "before_raw": before_raw})
        return response

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        snap = snapshot(instance, _SECTEUR_LABELS)
        nom = f"{instance.code} — {instance.nom}"
        log_action(request.user, "suppression_secteur",
                   detail=f"Secteur « {nom} » supprimé.",
                   meta={"snapshot": snap})
        return super().destroy(request, *args, **kwargs)

    def get_queryset(self):
        return (
            Secteur.objects.all()
            .annotate(
                production_total=Coalesce(
                    Sum("details_recolte__quantite",
                        filter=Q(details_recolte__ligne__fiche__statut="valide")),
                    0),
                last_recolte=Max(
                    "details_recolte__ligne__fiche__date",
                    filter=Q(details_recolte__ligne__fiche__statut="valide")),
            )
            .order_by("-id")
        )

    @action(detail=False, methods=["get"], url_path="export")
    def export(self, request):
        """Export Excel — liste des secteurs avec indicateurs."""
        year = request.query_params.get("year")
        year = int(year) if year else timezone.now().year

        wb = Workbook()
        ws = wb.active
        ws.title = "Secteurs"
        headers = ["Code", "Nom", "Superficie (ha)", "Statut", "Nb palmiers",
                   "Âge moyen plants", "Cible rendement (t/ha)",
                   "Production totale (régimes)", "Rendement (rég/ha)", "Dernière récolte"]
        _style_header(ws, headers)

        total_prod = 0
        for s in self.get_queryset().filter(statut="actif").order_by("code"):
            prod = int(getattr(s, "production_total", 0) or 0)
            sup = float(s.superficie_ha or 0)
            rend = round(prod / sup, 4) if sup else 0
            total_prod += prod
            ws.append([
                s.code, s.nom, float(s.superficie_ha),
                s.get_statut_display(),
                s.nb_palmiers or "",
                s.age_moyen_plants or "",
                float(s.rendement_cible_t_ha) if s.rendement_cible_t_ha else "",
                prod, rend,
                str(getattr(s, "last_recolte", None) or ""),
            ])

        total_row = ws.max_row + 1
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=8, value=total_prod).font = Font(bold=True)

        return _send_wb(wb, f"secteurs_export_{year}.xlsx")

    @action(detail=True, methods=["get"], url_path="analytics")
    def analytics(self, request, pk=None):
        from django.contrib.auth.models import User as DjangoUser
        secteur = self.get_object()
        today = timezone.now().date()
        year = int(request.query_params.get("year", today.year))
        prev_year = year - 1

        # Filtre optionnel par superviseur (created_by = User.pk)
        created_by_id = request.query_params.get("created_by")
        cb_user = None
        if created_by_id:
            try:
                cb_user = DjangoUser.objects.get(pk=int(created_by_id))
            except (ValueError, DjangoUser.DoesNotExist):
                pass
        detail_extra = {"ligne__fiche__created_by": cb_user} if cb_user else {}
        fiche_extra  = {"created_by": cb_user} if cb_user else {}
        pers_extra   = {"lignes_recolte__fiche__created_by": cb_user} if cb_user else {}

        month_labels = ["Jan", "Fev", "Mar", "Avr", "Mai", "Juin", "Juil", "Aout", "Sept", "Oct", "Nov", "Dec"]

        def monthly_totals(target_year):
            qs = (
                FicheRecolteDetail.objects.filter(secteur=secteur, ligne__fiche__date__year=target_year, ligne__fiche__statut="valide", **detail_extra)
                .values("ligne__fiche__date__month")
                .annotate(total=Sum("quantite"))
            )
            by_month = {row["ligne__fiche__date__month"]: int(row["total"] or 0) for row in qs}
            return {"labels": month_labels, "data": [by_month.get(m, 0) for m in range(1, 13)]}

        start_year = year - 4
        yearly_qs = (
            FicheRecolteDetail.objects.filter(secteur=secteur, ligne__fiche__date__year__gte=start_year, ligne__fiche__statut="valide", **detail_extra)
            .values("ligne__fiche__date__year")
            .annotate(total=Sum("quantite"))
            .order_by("ligne__fiche__date__year")
        )
        yearly_map = {r["ligne__fiche__date__year"]: int(r["total"] or 0) for r in yearly_qs}
        yearly_labels = list(range(start_year, year + 1))

        yearly_regime_qs = (
            FicheRecolteDetail.objects.filter(secteur=secteur, ligne__fiche__date__year__gte=start_year, ligne__fiche__statut="valide", **detail_extra)
            .values("ligne__fiche__date__year", "ligne__regime_type")
            .annotate(total=Sum("quantite"))
        )
        grands_by_year, moyens_by_year, petits_by_year = {}, {}, {}
        for r in yearly_regime_qs:
            yr = r["ligne__fiche__date__year"]
            t = int(r["total"] or 0)
            rt = r["ligne__regime_type"]
            if rt == "grands":
                grands_by_year[yr] = grands_by_year.get(yr, 0) + t
            elif rt == "moyens":
                moyens_by_year[yr] = moyens_by_year.get(yr, 0) + t
            elif rt == "petits":
                petits_by_year[yr] = petits_by_year.get(yr, 0) + t

        total_year = (
            FicheRecolteDetail.objects.filter(secteur=secteur, ligne__fiche__date__year=year, ligne__fiche__statut="valide", **detail_extra)
            .aggregate(total=Sum("quantite"))["total"] or 0
        )
        superficie = float(secteur.superficie_ha or 0)
        rendement_ha = round(float(total_year) / superficie, 4) if superficie else 0

        regime_qs = (
            FicheRecolteDetail.objects.filter(secteur=secteur, ligne__fiche__date__year=year, ligne__fiche__statut="valide", **detail_extra)
            .values("ligne__regime_type")
            .annotate(total=Sum("quantite"))
        )
        regime_map = {r["ligne__regime_type"]: int(r["total"] or 0) for r in regime_qs}

        nb_recolteurs_actifs = (
            Personnel.objects.filter(
                lignes_recolte__details__secteur=secteur,
                lignes_recolte__fiche__date__year=year,
                lignes_recolte__fiche__statut="valide",
                **pers_extra,
            )
            .distinct()
            .count()
        )

        top = (
            Personnel.objects.filter(
                lignes_recolte__details__secteur=secteur,
                lignes_recolte__fiche__date__year=year,
                lignes_recolte__fiche__statut="valide",
                **pers_extra,
            )
            .annotate(
                total=Coalesce(Sum("lignes_recolte__details__quantite",
                                   filter=Q(lignes_recolte__fiche__statut="valide")), 0),
                grands=Coalesce(Sum("lignes_recolte__details__quantite",
                                    filter=Q(lignes_recolte__regime_type="grands", lignes_recolte__fiche__statut="valide")), 0),
                moyens=Coalesce(Sum("lignes_recolte__details__quantite",
                                    filter=Q(lignes_recolte__regime_type="moyens", lignes_recolte__fiche__statut="valide")), 0),
                petits=Coalesce(Sum("lignes_recolte__details__quantite",
                                    filter=Q(lignes_recolte__regime_type="petits", lignes_recolte__fiche__statut="valide")), 0),
            )
            .values("id", "numero_telephone", "nom", "total", "grands", "moyens", "petits")
            .order_by("-total")[:10]
        )

        all_fiches = list(
            FicheRecolte.objects.filter(lignes__details__secteur=secteur, date__year=year, statut="valide", **fiche_extra)
            .order_by("date").distinct()
        )
        fiches_annee_rows = []
        for f in all_fiches:
            agg = (
                FicheRecolteDetail.objects.filter(secteur=secteur, ligne__fiche=f)
                .aggregate(
                    total=Sum("quantite"),
                    grands=Sum("quantite", filter=Q(ligne__regime_type="grands")),
                    moyens=Sum("quantite", filter=Q(ligne__regime_type="moyens")),
                    petits=Sum("quantite", filter=Q(ligne__regime_type="petits")),
                )
            )
            fiches_annee_rows.append({
                "fiche_id": f.id,
                "date": str(f.date),
                "total_regimes": int(agg["total"] or 0),
                "grands": int(agg["grands"] or 0),
                "moyens": int(agg["moyens"] or 0),
                "petits": int(agg["petits"] or 0),
            })
        last_rows = list(reversed(fiches_annee_rows[-10:]))

        return Response({
            "secteur": {
                "id": secteur.id, "code": secteur.code, "nom": secteur.nom,
                "superficie_ha": float(secteur.superficie_ha),
                "statut": secteur.statut,
                "nb_palmiers": secteur.nb_palmiers,
                "rendement_cible_t_ha": float(secteur.rendement_cible_t_ha) if secteur.rendement_cible_t_ha else None,
            },
            "filtre_superviseur": {
                "actif": cb_user is not None,
                "user_id": cb_user.id if cb_user else None,
                "nom_complet": f"{cb_user.first_name} {cb_user.last_name}".strip() or cb_user.username if cb_user else None,
            },
            "year": year,
            "monthly": {
                "current":  monthly_totals(year),
                "previous": monthly_totals(prev_year),
                "prev2":    monthly_totals(year - 2),
            },
            "yearly": {
                "labels": yearly_labels,
                "data":   [yearly_map.get(y, 0)    for y in yearly_labels],
                "grands": [grands_by_year.get(y, 0) for y in yearly_labels],
                "moyens": [moyens_by_year.get(y, 0) for y in yearly_labels],
                "petits": [petits_by_year.get(y, 0) for y in yearly_labels],
            },
            "stats": {
                "total_regimes": int(total_year),
                "rendement_ha": rendement_ha,
                "nb_recolteurs_actifs": nb_recolteurs_actifs,
                "regime_breakdown": {
                    "grands": regime_map.get("grands", 0),
                    "moyens": regime_map.get("moyens", 0),
                    "petits": regime_map.get("petits", 0),
                },
            },
            "top_recolteurs": list(top),
            "last_recoltes": last_rows,
            "fiches_annee": fiches_annee_rows,
        })

    @action(detail=True, methods=["get"], url_path="export")
    def export_details(self, request, pk=None):
        """Export Excel des récoltes pour un secteur."""
        secteur = self.get_object()
        today = timezone.now().date()
        year = int(request.query_params.get("year", today.year))

        wb = Workbook()
        ws = wb.active
        ws.title = f"Secteur {secteur.code}"
        headers = ["Date", "Téléphone récolteur", "Nom récolteur", "Type régime", "Quantité", "Fiche ID"]
        _style_header(ws, headers)

        qs = (
            FicheRecolteDetail.objects.select_related("ligne__fiche", "ligne__recolteur")
            .filter(secteur=secteur, ligne__fiche__date__year=year)
            .values("ligne__fiche__date", "ligne__recolteur__numero_telephone",
                    "ligne__recolteur__nom", "ligne__recolteur_nom",
                    "ligne__regime_type", "quantite", "ligne__fiche__id")
            .order_by("ligne__fiche__date")
        )

        total_qty = 0
        for row in qs:
            nom = row["ligne__recolteur__nom"] or row["ligne__recolteur_nom"] or ""
            qty = int(row["quantite"] or 0)
            total_qty += qty
            ws.append([str(row["ligne__fiche__date"]), row["ligne__recolteur__numero_telephone"] or "",
                        nom, row["ligne__regime_type"], qty, row["ligne__fiche__id"]])

        total_row = ws.max_row + 1
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=5, value=total_qty).font = Font(bold=True)

        return _send_wb(wb, f"secteur_{secteur.code}_recoltes_{year}.xlsx")

"""
travaux/views.py — API REST pour les fiches de travaux agricoles.

Expose FicheTravauxViewSet (ModelViewSet) :
  GET/POST   /api/travaux/           → liste et creation
  GET/PATCH  /api/travaux/:id/       → detail et modification partielle
  DELETE     /api/travaux/:id/       → suppression (interdit si valide)
  GET        /api/travaux/export/    → export Excel (admin uniquement)

Regles metier importantes :
  - Un admin ne peut pas creer de fiche travaux (action reservee aux superviseurs).
  - Une fiche validee ne peut etre modifiee que par l'admin.
  - La suppression d'une fiche validee est interdite a tous.
  - Chaque changement de statut declenche une notification au superviseur createur.
  - Toute modification est tracee dans ActionLog via log_action().
"""
import io

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response

from utils.permissions import IsAdmin
from utils.audit import log_action, snapshot, diff_fields
from accounts.utils import create_notification
from .models import FicheTravaux
from .serializers import FicheTravauxSerializer


def _send_wb(wb, filename: str) -> HttpResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def _style_header(ws, headers, fill_color="1F4E79"):
    fill = PatternFill(fill_type="solid", fgColor=fill_color)
    font = Font(bold=True, color="FFFFFF")
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = font; c.fill = fill; c.alignment = Alignment(horizontal="center")
    ws.auto_filter.ref = ws.dimensions


def _notify_statut_travaux(instance, old_statut, new_statut, actor=None):
    """Envoie une notification in-app au superviseur lors d'un changement de statut.

    Evite de notifier le superviseur d'une action qu'il a lui-meme effectuee
    (ex : il valide sa propre fiche). Seuls les changements vers 'valide'
    ou le rejet (brouillon depuis soumis/valide) declenchent une notification.
    """
    if not new_statut or new_statut == old_statut or not instance.created_by:
        return
    # Ne pas notifier le superviseur s'il est lui-meme l'acteur de l'action
    if actor and actor.pk == instance.created_by.pk:
        return
    ref = str(instance.periode_travaux) if instance.periode_travaux else f"#{instance.id}"
    if new_statut == "valide":
        create_notification(instance.created_by,
                            f"Votre fiche de travaux ({ref}) a ete validee.", "success",
                            "/travaux?tab=historique")
    elif new_statut == "brouillon" and old_statut in ("soumis", "valide"):
        create_notification(instance.created_by,
                            f"Votre fiche de travaux ({ref}) a ete rejetee.", "warning",
                            f"/mon-audit?action=rejet_travaux&fiche_id={instance.id}")


def _is_admin(user):
    try:
        return user.profile.is_admin
    except AttributeError:
        return False


_TRAVAUX_LABELS = {
    "superviseur_travaux":    "Superviseur",
    "nature_travaux":         "Nature des travaux",
    "periode_travaux":        "Période",
    "superficie_couverte_ha": "Superficie (ha)",
    "nb_personnes":           "Nombre de personnes",
    "observations":           "Observations",
    "statut":                 "Statut",
    "statut_avancement":      "Statut avancement",
}


class FicheTravauxViewSet(viewsets.ModelViewSet):
    serializer_class = FicheTravauxSerializer

    def get_queryset(self):
        from django.db.models import Q
        from agents.models import SuperviseurGeneral

        qs = FicheTravaux.objects.prefetch_related(
            "secteurs_couverts", "consommables", "repartitions"
        ).order_by("-id")
        if not _is_admin(self.request.user):
            q = Q(created_by=self.request.user)
            try:
                sup = SuperviseurGeneral.objects.get(user=self.request.user)
                if sup.nom:
                    q |= Q(superviseur_travaux__iexact=sup.nom)
            except SuperviseurGeneral.DoesNotExist:
                pass
            qs = qs.filter(q)
        return qs

    def perform_create(self, serializer):
        if _is_admin(self.request.user):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("L'administrateur ne peut pas créer de fiche travaux. Cette action est réservée aux superviseurs.")
        serializer.save(created_by=self.request.user)

    def create(self, request, *args, **kwargs):
        response = super().create(request, *args, **kwargs)
        if response.status_code in (200, 201):
            ref = response.data.get("periode_travaux") or f"#{response.data.get('id', '')}"
            snap = {
                "Superviseur": str(response.data.get("superviseur_travaux", "") or ""),
                "Nature des travaux": str(response.data.get("nature_travaux", "") or ""),
                "Période": str(response.data.get("periode_travaux", "") or ""),
                "Superficie (ha)": str(response.data.get("superficie_couverte_ha", "") or ""),
                "Nombre de personnes": str(response.data.get("nb_personnes", "") or ""),
                "Observations": str(response.data.get("observations", "") or ""),
            }
            log_action(request.user, "creation_travaux",
                       detail=f"Fiche travaux ({ref}) créée.",
                       meta={"snapshot": snap},
                       superviseur=request.user if not _is_admin(request.user) else None)
        return response

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        if instance.statut == "valide":
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("Une fiche travaux validée ne peut pas être supprimée.")
        if not _is_admin(request.user) and instance.created_by != request.user:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("Vous ne pouvez supprimer que vos propres fiches travaux.")
        ref = str(instance.periode_travaux) if instance.periode_travaux else f"#{instance.pk}"
        snap = {
            "Superviseur": instance.superviseur_travaux or "",
            "Nature des travaux": instance.get_nature_travaux_display() if instance.nature_travaux else "",
            "Période": instance.periode_travaux or "",
            "Superficie (ha)": str(instance.superficie_couverte_ha) if instance.superficie_couverte_ha is not None else "",
            "Statut": instance.get_statut_display(),
            "Nombre de personnes": str(instance.nb_personnes) if instance.nb_personnes is not None else "",
            "Observations": instance.observations or "",
        }
        log_action(request.user, "suppression_travaux",
                   detail=f"Fiche travaux ({ref}) supprimée.",
                   meta={"snapshot": snap},
                   superviseur=request.user if not _is_admin(request.user) else None)
        return super().destroy(request, *args, **kwargs)

    def get_permissions(self):
        if self.action in ("export",):
            return [IsAdmin()]
        return super().get_permissions()

    def _check_statut_transition(self, request, instance=None):
        """Verifie que la transition de statut demandee est autorisee.

        Regles :
          - Seul le createur peut passer une fiche a 'valide'.
          - Seul l'admin peut modifier une fiche deja validee.
          - Quand l'admin modifie une fiche validee, le motif est stocke
            sur l'instance (_audit_motif) pour etre inclus dans le log.

        Retourne None si la transition est autorisee, ou une Response d'erreur.
        """
        new_statut = request.data.get("statut")
        is_admin = _is_admin(request.user)
        is_creator = instance and instance.created_by_id == request.user.id
        if new_statut == "valide" and not is_creator:
            return Response(
                {"detail": "Seul le createur de la fiche peut la valider."},
                status=403,
            )
        if instance and instance.statut == "valide" and not is_admin:
            return Response({"detail": "Seul l'administrateur peut modifier une fiche validee."}, status=403)
        if instance and instance.statut == "valide" and is_admin:
            # Stocke le motif sur l'instance pour le recuperer dans le log d'audit
            instance._audit_user = request.user
            instance._audit_motif = request.data.get("motif", "")
        return None

    def partial_update(self, request, *args, **kwargs):
        instance = self.get_object()
        err = self._check_statut_transition(request, instance)
        if err:
            return err
        old_statut = instance.statut
        new_statut = request.data.get("statut")
        before = snapshot(instance, _TRAVAUX_LABELS)

        # Directly serialize + save to avoid calling self.update() which would double-log.
        serializer = self.get_serializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        if getattr(instance, "_prefetched_objects_cache", None):
            instance._prefetched_objects_cache = {}

        # Stamp or clear validation metadata
        if new_statut and new_statut != old_statut:
            if new_statut == "valide":
                instance.validated_by = request.user
                instance.validated_at = timezone.now()
                instance.save(update_fields=["validated_by", "validated_at"])
            elif new_statut == "brouillon" and old_statut in ("soumis", "valide"):
                instance.validated_by = None
                instance.validated_at = None
                instance.save(update_fields=["validated_by", "validated_at"])

        ref = str(instance.periode_travaux) if instance.periode_travaux else f"#{instance.pk}"
        instance.refresh_from_db()
        after = snapshot(instance, _TRAVAUX_LABELS)
        sup = instance.created_by

        if new_statut and new_statut != old_statut:
            transition_snap = {
                "Superviseur": instance.superviseur_travaux or "",
                "Nature des travaux": instance.get_nature_travaux_display() if instance.nature_travaux else "",
                "Période": instance.periode_travaux or "",
                "Statut précédent": old_statut,
            }
            if new_statut == "soumis":
                log_action(request.user, "soumission_travaux",
                           detail=f"Fiche travaux ({ref}) soumise pour validation.",
                           meta={"snapshot": transition_snap}, superviseur=sup)
            elif new_statut == "valide":
                log_action(request.user, "validation_travaux",
                           detail=f"Fiche travaux ({ref}) validée.",
                           meta={"snapshot": transition_snap}, superviseur=sup)
            elif new_statut == "brouillon" and old_statut in ("soumis", "valide"):
                motif = (request.data.get("motif") or getattr(instance, "_audit_motif", "") or "").strip()
                if motif:
                    transition_snap["Motif"] = motif
                detail_rejet = f"Fiche travaux ({ref}) rejetée."
                if motif:
                    detail_rejet += f" Motif : {motif}"
                log_action(request.user, "rejet_travaux",
                           detail=detail_rejet,
                           meta={"snapshot": transition_snap, "motif": motif}, superviseur=sup)
            else:
                changes = diff_fields(before, after)
                if changes:
                    log_action(request.user, "modification_travaux",
                               detail=f"Fiche travaux ({ref}) modifiée.",
                               meta={"changes": changes}, superviseur=sup)
        else:
            changes = diff_fields(before, after)
            if changes:
                log_action(request.user, "modification_travaux",
                           detail=f"Fiche travaux ({ref}) modifiée.",
                           meta={"changes": changes}, superviseur=sup)

        _notify_statut_travaux(instance, old_statut, new_statut, actor=request.user)
        return Response(serializer.data)

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        err = self._check_statut_transition(request, instance)
        if err:
            return err
        old_statut = instance.statut
        new_statut = request.data.get("statut")
        before = snapshot(instance, _TRAVAUX_LABELS)
        response = super().update(request, *args, **kwargs)
        if response.status_code == 200:
            ref = str(instance.periode_travaux) if instance.periode_travaux else f"#{instance.pk}"
            instance.refresh_from_db()
            after = snapshot(instance, _TRAVAUX_LABELS)
            sup = instance.created_by
            snap_base = {"Superviseur": instance.superviseur_travaux or "",
                         "Période": instance.periode_travaux or "",
                         "Statut précédent": old_statut}
            if new_statut and new_statut != old_statut and new_statut == "valide":
                log_action(request.user, "validation_travaux",
                           detail=f"Fiche travaux ({ref}) validée.",
                           meta={"snapshot": snap_base}, superviseur=sup)
            elif new_statut and new_statut != old_statut and new_statut == "brouillon" and old_statut in ("soumis", "valide"):
                motif_u = (request.data.get("motif") or "").strip()
                detail_u = f"Fiche travaux ({ref}) rejetée."
                if motif_u:
                    detail_u += f" Motif : {motif_u}"
                    snap_base["Motif"] = motif_u
                log_action(request.user, "rejet_travaux",
                           detail=detail_u,
                           meta={"snapshot": snap_base, "motif": motif_u}, superviseur=sup)
            else:
                changes = diff_fields(before, after)
                if changes:
                    log_action(request.user, "modification_travaux",
                               detail=f"Fiche travaux ({ref}) modifiée.",
                               meta={"changes": changes}, superviseur=sup)
        _notify_statut_travaux(instance, old_statut, new_statut, actor=request.user)
        return response

    @action(detail=False, methods=["get"], url_path="export")
    def export(self, request):
        """Export Excel des fiches de travaux."""
        year = request.query_params.get("year")
        today = timezone.now().date()
        year = int(year) if year else today.year

        wb = Workbook()
        ws_cons = wb.active
        ws_cons.title = "Consommables"
        ws_taches = wb.create_sheet("Tâches")

        headers_cons = ["Fiche ID", "Superviseur", "Nature travaux", "Superficie (ha)",
                         "Période", "Nb personnes", "Secteurs", "Désignation",
                         "Quantité", "Unité", "Prix unit.", "Prix total", "Fournisseur"]
        headers_taches = ["Fiche ID", "Superviseur", "Nature travaux", "Période",
                           "Secteurs", "Nom/Prénom", "Nature tâche", "Quantité",
                           "Prix unit.", "Salaire total", "Matricule"]
        _style_header(ws_cons, headers_cons)
        _style_header(ws_taches, headers_taches)

        total_cons = 0
        total_taches_sum = 0

        fiches = self.get_queryset().filter(created_at__year=year)
        for fiche in fiches:
            secteurs_codes = ", ".join(fiche.secteurs_couverts.values_list("code", flat=True))
            for c in fiche.consommables.all():
                prix = float((c.quantite or 0) * (c.prix_unitaire or 0))
                total_cons += prix
                ws_cons.append([
                    fiche.id, fiche.superviseur_travaux, fiche.nature_travaux,
                    float(fiche.superficie_couverte_ha or 0) or "",
                    fiche.periode_travaux, fiche.nb_personnes or "", secteurs_codes,
                    c.designation, float(c.quantite), c.unite, float(c.prix_unitaire), prix,
                    c.fournisseur or "",
                ])
            for r in fiche.repartitions.all():
                sal = float((r.quantite or 0) * (r.prix_unitaire or 0))
                total_taches_sum += sal
                ws_taches.append([
                    fiche.id, fiche.superviseur_travaux, fiche.nature_travaux,
                    fiche.periode_travaux, secteurs_codes,
                    r.nom_prenom, r.nature_taches, float(r.quantite),
                    float(r.prix_unitaire), sal, r.matricule_ouvrier or "",
                ])

        # Totaux
        tr = ws_cons.max_row + 1
        ws_cons.cell(tr, 1, "TOTAL").font = Font(bold=True)
        ws_cons.cell(tr, 12, round(total_cons, 2)).font = Font(bold=True)
        tr2 = ws_taches.max_row + 1
        ws_taches.cell(tr2, 1, "TOTAL").font = Font(bold=True)
        ws_taches.cell(tr2, 10, round(total_taches_sum, 2)).font = Font(bold=True)

        return _send_wb(wb, f"travaux_export_{year}.xlsx")
